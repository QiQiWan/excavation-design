from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from statistics import mean
from typing import Any

from openpyxl import load_workbook

from app.schemas.domain import Borehole, BoreholeLayer, GroundwaterRecord, SoilParameters, Stratum

REQUIRED_COLUMNS = [
    "borehole_code",
    "x",
    "y",
    "collar_elevation",
    "borehole_depth",
    "layer_index",
    "stratum_code",
    "stratum_name",
    "top_depth",
    "bottom_depth",
]

OPTIONAL_COLUMNS = [
    "unit_weight",
    "saturated_unit_weight",
    "cohesion",
    "friction_angle",
    "elastic_modulus",
    "compression_modulus",
    "poisson_ratio",
    "permeability",
    "permeability_x",
    "permeability_y",
    "permeability_z",
    "k0",
    "horizontal_subgrade_modulus",
    "water_level",
]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


@dataclass
class BoreholeImportResult:
    success: bool
    borehole_count: int = 0
    layer_count: int = 0
    stratum_count: int = 0
    boreholes: list[Borehole] = field(default_factory=list)
    strata: list[Stratum] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def as_response(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "boreholeCount": self.borehole_count,
            "layerCount": self.layer_count,
            "stratumCount": self.stratum_count,
            "warnings": self.warnings,
            "errors": self.errors,
            "boreholes": [b.model_dump(mode="json", by_alias=True) for b in self.boreholes],
            "strata": [s.model_dump(mode="json", by_alias=True) for s in self.strata],
        }


def _clean_header(name: str) -> str:
    return str(name or "").strip().lower()


def _parse_float(row: dict[str, str], key: str, row_no: int, errors: list[str], required: bool = False) -> float | None:
    raw = row.get(key, "")
    if raw is None or str(raw).strip() == "":
        if required:
            errors.append(f"第 {row_no} 行：{key} 不能为空。")
        return None
    try:
        return float(str(raw).strip())
    except ValueError:
        errors.append(f"第 {row_no} 行：{key} 必须为数值，实际为 {raw!r}。")
        return None


def read_csv_bytes(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        return []
    reader.fieldnames = [_clean_header(field) for field in reader.fieldnames]
    return [{_clean_header(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items()} for row in reader]


def read_excel_bytes(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_header(h) for h in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        result.append({headers[i]: "" if value is None else str(value) for i, value in enumerate(row) if i < len(headers)})
    return result


def _warn_abnormal_parameters(row: dict[str, str], row_no: int, warnings: list[str]) -> None:
    checks = {
        "unit_weight": (5.0, 30.0, "kN/m3"),
        "saturated_unit_weight": (5.0, 35.0, "kN/m3"),
        "cohesion": (0.0, 500.0, "kPa"),
        "friction_angle": (0.0, 50.0, "degree"),
        "elastic_modulus": (0.1, 500.0, "MPa"),
        "compression_modulus": (0.1, 1000.0, "MPa"),
        "poisson_ratio": (0.0, 0.5, ""),
        "permeability": (0.0, 1.0, "m/s"),
        "permeability_x": (0.0, 1.0, "m/s"),
        "permeability_y": (0.0, 1.0, "m/s"),
        "permeability_z": (0.0, 1.0, "m/s"),
        "k0": (0.0, 3.0, ""),
        "horizontal_subgrade_modulus": (1.0, 1.0e7, "kN/m3"),
    }
    for key, (low, high, unit) in checks.items():
        raw = row.get(key)
        if raw is None or str(raw).strip() == "":
            continue
        try:
            value = float(raw)
        except ValueError:
            continue
        if value < low or value > high:
            suffix = f" {unit}" if unit else ""
            warnings.append(f"第 {row_no} 行：{key}={value}{suffix} 超出 MVP 常规范围 [{low}, {high}]，已保留但需复核。")




def _compact_warnings(warnings: list[str]) -> list[str]:
    """Collapse repeated row-level range warnings from repeated strata.

    Real projects often repeat one stratum in many boreholes. Showing the same
    parameter warning once per row obscures the actual number of distinct risks.
    """
    grouped: dict[str, tuple[str, int]] = {}
    ordered: list[str] = []
    for warning in warnings:
        match = re.match(r"第 (\d+) 行：(.*超出 MVP 常规范围.*)", warning)
        if not match:
            if warning not in ordered:
                ordered.append(warning)
            continue
        row_no, body = match.groups()
        if body not in grouped:
            grouped[body] = (row_no, 1)
        else:
            first, count = grouped[body]
            grouped[body] = (first, count + 1)
    for body, (first, count) in grouped.items():
        suffix = f"（共 {count} 行，首见第 {first} 行）" if count > 1 else f"（第 {first} 行）"
        ordered.append(f"{body}{suffix}")
    return ordered


def _merge_soil_parameters(rows: list[dict[str, str]]) -> SoilParameters:
    def avg(key: str) -> float | None:
        values: list[float] = []
        for row in rows:
            raw = row.get(key)
            if raw is None or str(raw).strip() == "":
                continue
            try:
                values.append(float(raw))
            except ValueError:
                continue
        return round(mean(values), 6) if values else None

    permeability = avg("permeability")
    permeability_x = avg("permeability_x")
    permeability_y = avg("permeability_y")
    permeability_z = avg("permeability_z")
    return SoilParameters(
        unit_weight=avg("unit_weight"),
        saturated_unit_weight=avg("saturated_unit_weight"),
        cohesion=avg("cohesion"),
        friction_angle=avg("friction_angle"),
        elastic_modulus=avg("elastic_modulus"),
        compression_modulus=avg("compression_modulus"),
        poisson_ratio=avg("poisson_ratio"),
        permeability_x=permeability_x if permeability_x is not None else permeability,
        permeability_y=permeability_y if permeability_y is not None else permeability,
        permeability_z=permeability_z if permeability_z is not None else permeability,
        k0=avg("k0"),
        horizontal_subgrade_modulus=avg("horizontal_subgrade_modulus"),
    )


def parse_borehole_rows(rows: list[dict[str, str]], source_file: str | None = None) -> BoreholeImportResult:
    warnings: list[str] = []
    errors: list[str] = []
    if not rows:
        return BoreholeImportResult(success=False, errors=["导入文件为空或没有表头。"])

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in rows[0]]
    if missing_columns:
        return BoreholeImportResult(success=False, errors=[f"缺少必填列：{', '.join(missing_columns)}"])

    bh_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    stratum_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    stratum_names: dict[str, str] = {}

    for index, row in enumerate(rows, start=2):
        code = str(row.get("borehole_code", "")).strip()
        if not code:
            errors.append(f"第 {index} 行：borehole_code 不能为空。")
            continue
        for key in ["x", "y", "collar_elevation", "borehole_depth", "top_depth", "bottom_depth"]:
            _parse_float(row, key, index, errors, required=True)
        top = _parse_float(row, "top_depth", index, errors)
        bottom = _parse_float(row, "bottom_depth", index, errors)
        depth = _parse_float(row, "borehole_depth", index, errors)
        if top is not None and bottom is not None and top >= bottom:
            errors.append(f"第 {index} 行：top_depth 必须小于 bottom_depth。")
        if bottom is not None and depth is not None and bottom > depth + 1e-9:
            errors.append(f"第 {index} 行：bottom_depth 不得大于 borehole_depth。")
        stratum_code = str(row.get("stratum_code", "")).strip()
        stratum_name = str(row.get("stratum_name", "")).strip()
        if not stratum_code:
            errors.append(f"第 {index} 行：stratum_code 不能为空。")
        if not stratum_name:
            errors.append(f"第 {index} 行：stratum_name 不能为空。")
        if stratum_code:
            previous = stratum_names.get(stratum_code)
            if previous is not None and stratum_name and previous != stratum_name:
                warnings.append(f"第 {index} 行：同一地层编号 {stratum_code} 出现不同名称：{previous} / {stratum_name}。")
            elif stratum_name:
                stratum_names[stratum_code] = stratum_name
            stratum_rows[stratum_code].append(row)
        _warn_abnormal_parameters(row, index, warnings)
        bh_rows[code].append(row)

    if errors:
        return BoreholeImportResult(success=False, warnings=_compact_warnings(warnings), errors=errors)

    boreholes: list[Borehole] = []
    layer_count = 0
    for code, group in bh_rows.items():
        sorted_group = sorted(group, key=lambda r: float(r.get("top_depth") or 0.0))
        first = sorted_group[0]
        collar = float(first["collar_elevation"])
        depth = float(first["borehole_depth"])
        x = float(first["x"])
        y = float(first["y"])
        layers: list[BoreholeLayer] = []
        previous_bottom: float | None = None
        for row in sorted_group:
            top = float(row["top_depth"])
            bottom = float(row["bottom_depth"])
            if previous_bottom is not None and top < previous_bottom - 1e-9:
                errors.append(f"钻孔 {code} 层序交叉：top_depth={top} 小于上一层 bottom_depth={previous_bottom}。")
            previous_bottom = bottom
            layers.append(
                BoreholeLayer(
                    stratum_code=str(row["stratum_code"]).strip(),
                    stratum_name=str(row["stratum_name"]).strip(),
                    top_depth=top,
                    bottom_depth=bottom,
                    top_elevation=round(collar - top, 6),
                    bottom_elevation=round(collar - bottom, 6),
                )
            )
            layer_count += 1
        water_levels: list[GroundwaterRecord] = []
        water_raw = str(first.get("water_level", "")).strip()
        if water_raw:
            try:
                water_levels.append(GroundwaterRecord(water_level=collar - float(water_raw), description="Imported as elevation from depth"))
            except ValueError:
                warnings.append(f"钻孔 {code} water_level 无法解析，已忽略。")
        boreholes.append(Borehole(code=code, x=x, y=y, collar_elevation=collar, depth=depth, layers=layers, water_levels=water_levels, source_file=source_file))

    if errors:
        return BoreholeImportResult(success=False, warnings=_compact_warnings(warnings), errors=errors)

    palette = ["#b7c9a8", "#e2c290", "#b6d7e8", "#d5b8d8", "#f4b183", "#a9d18e", "#c9c9c9"]
    strata: list[Stratum] = []
    for idx, (code, rows_for_stratum) in enumerate(sorted(stratum_rows.items())):
        name = stratum_names.get(code) or str(rows_for_stratum[0].get("stratum_name") or code)
        strata.append(
            Stratum(
                code=code,
                name=name,
                color=palette[idx % len(palette)],
                parameters=_merge_soil_parameters(rows_for_stratum),
                parameter_source="imported",
                confidence="medium",
            )
        )

    return BoreholeImportResult(
        success=True,
        borehole_count=len(boreholes),
        layer_count=layer_count,
        stratum_count=len(strata),
        boreholes=boreholes,
        strata=strata,
        warnings=_compact_warnings(warnings),
        errors=[],
    )
