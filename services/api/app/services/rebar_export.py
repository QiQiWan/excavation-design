from __future__ import annotations

import csv
import json
import shutil
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.domain import Project
from app.services.rebar_detailing import build_rebar_detailing
from app.version import EXPORT_SCHEMA_VERSION, SOFTWARE_VERSION


WORKBOOK_ROW_LIMIT = 5000


def _json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _columns(rows: Iterable[dict[str, Any]]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                ordered.append(key)
                seen.add(key)
    return ordered


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = _columns(rows)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=cols or ["message"])
        writer.writeheader()
        if rows:
            for row in rows:
                writer.writerow({key: _json_text(row.get(key)) for key in cols})
        else:
            writer.writerow({"message": "no records"})


def _safe_sheet_name(value: str) -> str:
    invalid = set('[]:*?/\\')
    name = "".join("_" if ch in invalid else ch for ch in value).strip() or "Sheet"
    return name[:31]


def _write_workbook(path: Path, tables: list[tuple[str, list[dict[str, Any]]]]) -> None:
    """Write large reinforcement schedules in streaming mode.

    Normal openpyxl worksheets retain every cell object in memory and become
    disproportionately slow once individual-bar tables exceed tens of thousands
    of cells.  Write-only mode keeps the human workbook bounded while the CSV
    and JSON files remain the complete machine-readable sources.
    """
    wb = Workbook(write_only=True)
    for title, rows in tables:
        ws = wb.create_sheet(_safe_sheet_name(title))
        cols = _columns(rows)
        if not cols:
            ws.append(["message"])
            ws.append(["no records"])
            continue
        for index, key in enumerate(cols, start=1):
            sample = [len(str(key))] + [len(_json_text(row.get(key))) for row in rows[:100]]
            ws.column_dimensions[get_column_letter(index)].width = min(max(max(sample, default=8) + 2, 10), 36)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(len(rows) + 1, 2)}"
        header = []
        for key in cols:
            cell = WriteOnlyCell(ws, value=key)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header.append(cell)
        ws.append(header)
        for row in rows:
            ws.append([_json_text(row.get(key)) for key in cols])
    wb.save(path)


def _summary_rows(detailing: dict[str, Any]) -> list[dict[str, Any]]:
    summary = detailing.get("summary") or {}
    return [{"metric": key, "value": _json_text(value)} for key, value in summary.items()]


def _write_json_object_with_array(path: Path, metadata: dict[str, Any], array_key: str, rows: list[dict[str, Any]]) -> None:
    """Stream a large JSON array without constructing a second serialized copy."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        prefix = json.dumps(metadata, ensure_ascii=False, separators=(",", ":"))
        handle.write(prefix[:-1] + f',"{array_key}":[')
        for index, row in enumerate(rows):
            if index:
                handle.write(",")
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")))
        handle.write("]}")


def _compact_machine_payload(detailing: dict[str, Any]) -> dict[str, Any]:
    external_keys = {
        "individualBars": "00_machine_data/individual_rebar_geometry.json",
        "fabricationSegments": "10_schedules/fabrication_segments.csv",
        "fabricationBbs": "10_schedules/fabrication_bbs.csv",
        "fabricationSplices": "10_schedules/splice_schedule.csv",
        "geometricSpacingChecks": "20_checks/spacing_checks.csv",
    }
    compact = {key: value for key, value in detailing.items() if key not in external_keys}
    fabrication = dict(compact.get("fabrication") or {})
    for key in ("fabricationSegments", "barBendingSchedule", "spliceRecords", "geometricSpacingChecks"):
        fabrication.pop(key, None)
    if fabrication:
        compact["fabrication"] = fabrication
    compact["externalizedCompleteData"] = external_keys
    compact["dataCompleteness"] = "complete_across_compact_json_streamed_geometry_and_csv_tables"
    return compact


def _zip_tree_fast(source_root: Path, final_zip: Path) -> None:
    if final_zip.exists():
        final_zip.unlink()
    with zipfile.ZipFile(final_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1, allowZip64=True) as archive:
        for file in sorted(source_root.rglob("*")):
            if file.is_file():
                archive.write(file, file.relative_to(source_root.parent).as_posix())


def export_rebar_detailing_package(project: Project, output_dir: Path, mode: str = "balanced") -> Path:
    """Export a human-usable reinforcement detailing package.

    JSON remains the lossless machine interchange file. CSV/XLSX tables provide direct
    review, quantity take-off and fabrication hand-off, while the README explains which
    files are suitable for design review, CAD coordination and downstream automation.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    detailing = build_rebar_detailing(project, mode=mode)
    package_name = f"{project.id}_rebar_detailing_package_v{SOFTWARE_VERSION.replace('.', '_')}"
    final_zip = output_dir / f"{package_name}.zip"

    with tempfile.TemporaryDirectory(prefix="pitguard_rebar_") as temp:
        root = Path(temp) / package_name
        data_dir = root / "00_machine_data"
        schedules_dir = root / "10_schedules"
        checks_dir = root / "20_checks"
        guidance_dir = root / "90_guidance"
        for folder in (data_dir, schedules_dir, checks_dir, guidance_dir):
            folder.mkdir(parents=True, exist_ok=True)

        compact_payload = _compact_machine_payload(detailing)
        (data_dir / "rebar_detailing_full.json").write_text(
            json.dumps(compact_payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )
        _write_json_object_with_array(
            data_dir / "individual_rebar_geometry.json",
            {"projectId": project.id, "schemaVersion": EXPORT_SCHEMA_VERSION, "barCount": len(detailing.get("individualBars", []))},
            "bars",
            list(detailing.get("individualBars", [])),
        )
        (data_dir / "rebar_design_scheme.json").write_text(
            json.dumps(detailing.get("designScheme", {}), ensure_ascii=False, indent=2), encoding="utf-8"
        )

        table_specs: list[tuple[str, str, list[dict[str, Any]]]] = [
            ("summary", "钢筋包汇总", _summary_rows(detailing)),
            ("rebar_mark_schedule", "钢筋编号表", list(detailing.get("entries") or [])),
            ("individual_bars", "逐根钢筋几何", list(detailing.get("individualBars") or [])),
            ("fabrication_bbs", "加工下料表BBS", list(detailing.get("fabricationBbs") or [])),
            ("fabrication_segments", "加工分段表", list(detailing.get("fabricationSegments") or [])),
            ("splice_schedule", "接头与套筒表", list(detailing.get("fabricationSplices") or detailing.get("spliceSchedule") or [])),
            ("cage_segments", "钢筋笼分段表", list(detailing.get("cageSegments") or [])),
            ("lifting_plan", "吊装计划", list(detailing.get("liftingPlan") or [])),
            ("construction_joint_plan", "施工缝计划", list(detailing.get("constructionJointPlan") or [])),
        ]
        check_specs: list[tuple[str, str, list[dict[str, Any]]]] = [
            ("spacing_checks", "净距检查", list(detailing.get("geometricSpacingChecks") or [])),
            ("cover_conflict_checks", "保护层冲突检查", list(detailing.get("coverConflictChecks") or [])),
            ("bend_radius_checks", "弯曲半径检查", list(detailing.get("bendRadiusChecks") or [])),
            ("signoff_checklist", "签审检查表", list(detailing.get("signoffChecklist") or [])),
        ]
        for filename, _, rows in table_specs:
            _write_csv(schedules_dir / f"{filename}.csv", rows)
        for filename, _, rows in check_specs:
            _write_csv(checks_dir / f"{filename}.csv", rows)

        workbook_tables: list[tuple[str, list[dict[str, Any]]]] = []
        workbook_truncation: list[dict[str, Any]] = []
        for filename, title, rows in table_specs + check_specs:
            workbook_rows = rows[:WORKBOOK_ROW_LIMIT]
            workbook_tables.append((title, workbook_rows))
            if len(rows) > len(workbook_rows):
                workbook_truncation.append({
                    "table": filename,
                    "totalRows": len(rows),
                    "workbookRows": len(workbook_rows),
                    "completeSource": f"10_schedules/{filename}.csv" if (schedules_dir / f"{filename}.csv").exists() else f"20_checks/{filename}.csv",
                })
        _write_workbook(root / "rebar_detailing_schedules.xlsx", workbook_tables)

        latest = project.calculation_results[-1] if project.calculation_results else None
        calculation_assurance = dict(getattr(latest, "calculation_assurance", {}) or {}) if latest else {}
        manifest = {
            "schemaVersion": EXPORT_SCHEMA_VERSION,
            "softwareVersion": SOFTWARE_VERSION,
            "projectId": project.id,
            "projectName": project.name,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "mode": mode,
            "packageType": "rebar_detailing_zip",
            "humanReadablePrimary": "rebar_detailing_schedules.xlsx",
            "machineReadablePrimary": "00_machine_data/rebar_detailing_full.json + 00_machine_data/individual_rebar_geometry.json",
            "machineDataLayout": "large arrays externalized to streamed geometry JSON and complete CSV schedules",
            "cadDrawingSource": "Use the separate CAD drawing package (scope=rebar or full) for DXF construction drawings.",
            "workbookRowLimitPerTable": WORKBOOK_ROW_LIMIT,
            "workbookTruncation": workbook_truncation,
            "completeTabularSources": "All CSV files are complete; JSON files retain full geometry and semantics.",
            "summary": detailing.get("summary", {}),
            "calculationBaseline": {
                "calculationResultId": getattr(latest, "id", None),
                "calculationContractId": getattr(latest, "calculation_contract_id", None),
                "inputSnapshotHash": getattr(latest, "input_snapshot_hash", None),
                "adoptedDesignSnapshotHash": getattr(latest, "adopted_design_snapshot_hash", None),
                "resultHash": getattr(latest, "result_hash", None),
                "assuranceStatus": calculation_assurance.get("status"),
            },
        }
        (root / "package_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

        readme = f"""# PitGuard 钢筋加工深化包使用说明

项目：{project.name}  
项目 ID：{project.id}  
软件版本：{SOFTWARE_VERSION}  
配筋模式：{mode}

## 为什么包内仍保留 JSON

JSON 是逐根钢筋几何、构件关联、规则检查、套筒、吊装和碰撞信息的无损机器交换格式，适合二次开发、BIM/数字化加工接口和追溯。大型项目采用紧凑主 JSON + 流式逐根几何 JSON + 完整 CSV，避免同一批逐根数据重复存储和长时间压缩。JSON 本身不等同于施工图，也不适合作为现场人员的主要阅读文件。

## 推荐使用顺序

1. 打开根目录 `rebar_detailing_schedules.xlsx`，进行钢筋编号、下料、接头、钢筋笼分段、吊装和检查项的人工复核。
2. `10_schedules/*.csv` 可导入 Excel、ERP、钢筋翻样或加工设备中间系统；导入前需按设备字段映射确认单位。大型项目中 XLSX 每张表最多显示 {WORKBOOK_ROW_LIMIT} 行，CSV 与 JSON 保留完整记录，截断信息记录在 `package_manifest.json`。
3. `20_checks/*.csv` 用于处理净距、保护层、弯曲半径和签审问题。
4. `00_machine_data/rebar_detailing_full.json` 保存紧凑语义与外部数据索引；`individual_rebar_geometry.json` 和各 CSV 保存完整逐根与加工数据。
5. 需要可打印、可审签的钢筋施工图时，请在系统中下载“CAD 图纸包（钢筋范围）”或“正式图纸发行包”，其中包含 DXF/PDF/图纸目录和审签信息。

## 单位

长度字段按字段名区分：`Mm` 为毫米，`M` 为米；重量为 kg；面积按字段说明为 mm² 或 mm²/m。逐根钢筋坐标沿用项目坐标系，坐标值单位为米。

## 工程边界

自动下料、锚固、搭接、机械连接、钢筋笼分段、吊点与碰撞检查属于深化辅助结果。正式加工和施工前仍需结合企业标准图集、材料复验、接头工艺评定、吊装专项方案、施工缝位置和现场条件完成专业复核与签审。
"""
        (guidance_dir / "README_USAGE.md").write_text(readme, encoding="utf-8")
        (root / "README.txt").write_text(readme.replace("# ", "").replace("## ", "\n"), encoding="utf-8")

        _zip_tree_fast(root, final_zip)
    return final_zip
